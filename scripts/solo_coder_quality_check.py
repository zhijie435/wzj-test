#!/usr/bin/env python3
"""
Solo Coder spreadsheet quality checker.

It learns batch context from the workbook and checks a new row against the
rules observed in the "质检备注" column and guideline sheets.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from collections import Counter
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - runtime guidance
    load_workbook = None


DEFAULT_WORKBOOK = "/Users/wuzhijie/Downloads/成都Solo Coder 0601.xlsx"
DATA_SHEET = "数据总表"
REMARK_FIELD = "质检备注"

REQUIRED_FIELDS = [
    "Repo ID",
    "Repo URL",
    "Commit ID",
    "Trae Session ID",
    "User Prompt",
    "任务类型",
    "业务领域",
    "修改范围",
    "任务难度",
    "任务是否完成",
    "过程与产物是否满意",
    "做题人",
    "提交日期",
    "是否提交字节",
]

TASK_TYPES = {"0-1代码生成", "Feature迭代", "Bug修复", "代码理解", "代码重构", "其他"}
CHANGE_SCOPES = {"无需修改", "单文件", "跨模块多文件", "跨系统多模块"}

HARD = "打回"
DISCARD = "废弃"
REVIEW = "人工复核"


@dataclass
class Issue:
    rule: str
    field: str
    severity: str
    message: str
    suggestion: str = ""
    evidence: str = ""


def text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def compact(value: Any) -> str:
    value = unicodedata.normalize("NFKC", text(value)).lower()
    return re.sub(r"[\s\W_]+", "", value, flags=re.UNICODE)


def trigrams(value: str) -> set[str]:
    value = compact(value)
    if len(value) < 3:
        return {value} if value else set()
    return {value[i : i + 3] for i in range(len(value) - 2)}


def lcs_len(a: str, b: str, limit: int = 400) -> int:
    a = compact(a)[:limit]
    b = compact(b)[:limit]
    if not a or not b:
        return 0
    if len(a) > len(b):
        a, b = b, a
    prev = [0] * (len(a) + 1)
    best = 0
    for ch_b in b:
        cur = [0] * (len(a) + 1)
        for i, ch_a in enumerate(a, 1):
            if ch_a == ch_b:
                cur[i] = prev[i - 1] + 1
                if cur[i] > best:
                    best = cur[i]
        prev = cur
    return best


def jaccard(a: str, b: str) -> float:
    left = trigrams(a)
    right = trigrams(b)
    if not left or not right:
        return 0.0
    return len(left & right) / len(left | right)


def containment(a: str, b: str) -> float:
    left = trigrams(a)
    right = trigrams(b)
    if not left or not right:
        return 0.0
    return len(left & right) / min(len(left), len(right))


def load_rows(workbook_path: str, sheet_name: str = DATA_SHEET) -> list[dict[str, Any]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is required to read .xlsx files")
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    header = [text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows: list[dict[str, Any]] = []
    for cells in ws.iter_rows(min_row=2):
        row = {header[i]: cells[i].value for i in range(min(len(header), len(cells)))}
        if any(text(v) for v in row.values()):
            rows.append(row)
    return rows


def workbook_summary(rows: list[dict[str, Any]]) -> dict[str, Any]:
    remarks = [text(row.get(REMARK_FIELD)) for row in rows if text(row.get(REMARK_FIELD))]
    rule_pattern = re.compile(r"❌\s*规则\s*([^：:]+)[：:]\s*([^\n]+)")
    rules: Counter[str] = Counter()
    for remark in remarks:
        for code, title in rule_pattern.findall(remark):
            rules[f"规则 {code.strip()}：{title.strip()}"] += 1
    return {
        "data_rows": len(rows),
        "non_empty_quality_notes": len(remarks),
        "quality_status": Counter(text(row.get("质检状态")) or "空" for row in rows),
        "satisfaction": Counter(text(row.get("过程与产物是否满意")) or "空" for row in rows),
        "top_quality_rules": rules.most_common(25),
    }


def is_blank(value: Any) -> bool:
    return text(value) == ""


def add(issues: list[Issue], rule: str, field: str, severity: str, message: str, suggestion: str = "", evidence: str = "") -> None:
    issues.append(Issue(rule, field, severity, message, suggestion, evidence))


def session_key(value: str) -> str:
    value = text(value)
    match = re.search(r":([0-9a-fA-F]{16,}(?:_[0-9a-fA-F]{8,})?)", value)
    if match:
        return match.group(1).lower()
    return compact(value)


def split_reason(reason: str) -> tuple[str, str]:
    product = ""
    process = ""
    product_match = re.search(r"(?:产物|结果)不满意[：:](.*?)(?=过程不满意[：:]|$)", reason, re.S)
    process_match = re.search(r"过程不满意[：:](.*)$", reason, re.S)
    if product_match:
        product = product_match.group(1).strip()
    if process_match:
        process = process_match.group(1).strip()
    return product, process


def check_required(row: dict[str, Any], issues: list[Issue]) -> None:
    for field in REQUIRED_FIELDS:
        if is_blank(row.get(field)):
            add(issues, "规则 3", field, HARD, f"{field} 为空。", "补齐该必填字段。")


def check_repo_fields(row: dict[str, Any], issues: list[Issue]) -> None:
    repo_id = text(row.get("Repo ID"))
    if repo_id and not re.match(r"^(?:[a-z]{1,8}-\d+-\d+|[AB]-\d+-\d+)$", repo_id):
        add(
            issues,
            "规则 1",
            "Repo ID",
            HARD,
            "Repo ID 格式不符合规范。",
            "自建仓库用“姓名拼音缩写-Repo编号-题目编号”，题库仓库用“A/B-编号-题目编号”。",
            repo_id,
        )

    repo_url = text(row.get("Repo URL"))
    if repo_url and not re.match(r"^https://github\.com/[^/\s]+/[^/\s]+/?$", repo_url):
        add(
            issues,
            "规则 32",
            "Repo URL",
            HARD,
            "Repo URL 应为 GitHub 仓库主页，不能是提交页、分支页或其他链接。",
            "填写形如 https://github.com/owner/repo 的公开仓库地址。",
            repo_url,
        )

    commit_id = text(row.get("Commit ID"))
    if commit_id and not re.match(r"^[0-9a-fA-F]{40,}$", commit_id):
        add(
            issues,
            "规则 33",
            "Commit ID",
            HARD,
            "Commit ID 不是完整的 Git 提交哈希。",
            "复制 40 位以上的完整 commit hash，不要填短哈希或 URL。",
            commit_id,
        )

    session = text(row.get("Trae Session ID"))
    if session:
        looks_like_prompt = len(session) > 80 and re.search(r"[\u4e00-\u9fff]", session) and not re.search(r"Trae CN\.T|:[0-9a-fA-F]{16,}", session)
        valid_session = "Trae CN.T" in session or re.search(r"^\.\d+:[0-9a-fA-F]{16,}_[0-9a-fA-F]{8,}", session)
        if looks_like_prompt or not valid_session:
            add(
                issues,
                "规则 2",
                "Trae Session ID",
                HARD,
                "Trae Session ID 格式异常，可能填成了提示词或非会话 ID。",
                "填写 Trae 导出的 Session ID；不要把 User Prompt 粘到这一列。",
                session[:160],
            )


def check_task_consistency(row: dict[str, Any], issues: list[Issue]) -> None:
    task_type = text(row.get("任务类型"))
    prompt = text(row.get("User Prompt"))
    scope = text(row.get("修改范围"))
    difficulty = text(row.get("任务难度"))

    if task_type and task_type not in TASK_TYPES:
        add(issues, "规则 14", "任务类型", HARD, "任务类型不是规范枚举值。", f"可选值：{', '.join(sorted(TASK_TYPES))}。", task_type)
    if scope and scope not in CHANGE_SCOPES:
        add(issues, "规则 3", "修改范围", HARD, "修改范围不是规范枚举值。", f"可选值：{', '.join(sorted(CHANGE_SCOPES))}。", scope)

    if task_type == "Bug修复" and prompt and not re.search(r"bug|修复|问题|报错|异常|失败|错误|无法|崩溃|不生效", prompt, re.I):
        add(
            issues,
            "规则 14",
            "任务类型",
            REVIEW,
            "任务类型选择了 Bug修复，但 User Prompt 没有明显的问题/修复指向。",
            "若确实是修复题，在提示词中明确“修复 xxx 问题”；否则调整任务类型。",
            prompt[:160],
        )

    if task_type == "代码理解":
        if scope and scope != "无需修改":
            add(issues, "规则 20", "修改范围", DISCARD, "代码理解任务的修改范围应为“无需修改”。", "改为“无需修改”，或将任务类型改为代码生成/迭代/重构。", scope)
        if difficulty == "简单":
            add(issues, "规则 22", "任务难度", DISCARD, "简单难度的代码理解任务按规则需要废弃。", "提高任务难度或不要提交该题。")
        if prompt and re.search(r"实现|新增|开发|修改|修复|生成|创建|接入|优化", prompt):
            add(
                issues,
                "规则 14",
                "User Prompt",
                HARD,
                "代码理解任务的提示词包含明显实现/修改动作。",
                "代码理解题应偏阅读、解释、梳理、分析，不应要求改代码。",
                prompt[:160],
            )


def check_satisfaction_and_reason(row: dict[str, Any], issues: list[Issue]) -> None:
    done = text(row.get("任务是否完成"))
    satisfaction = text(row.get("过程与产物是否满意"))
    reason = text(row.get("不满意原因"))
    prompt = text(row.get("User Prompt"))

    if satisfaction == "不满意" and not reason:
        add(issues, "规则 4", "不满意原因", HARD, "不满意时必须填写不满意原因。", "补写产物不满意和过程不满意，且不少于 30 字。")
        return
    if satisfaction == "满意" and reason:
        add(issues, "规则 4", "不满意原因", HARD, "满意时不应填写不满意原因。", "清空不满意原因，或把满意度改为不满意。", reason[:160])
    if done == "未完成任务" and satisfaction == "满意":
        add(issues, "规则 11", "过程与产物是否满意", HARD, "任务未完成不能选择满意。", "改为不满意，并说明未完成的产物和过程问题。")

    if not reason:
        return

    if len(compact(reason)) < 30:
        add(issues, "规则 5", "不满意原因", HARD, "不满意原因字数不足 30 字。", "补充触发节点、实际行为、业务影响。", reason)

    product, process = split_reason(reason)
    has_product_marker = bool(re.search(r"(?:产物|结果)不满意[：:]", reason))
    has_process_marker = bool(re.search(r"过程不满意[：:]", reason))

    if done == "未完成任务" and (not has_product_marker or not has_process_marker):
        add(
            issues,
            "规则 37",
            "不满意原因",
            HARD,
            "未完成任务的不满意原因需要同时包含“产物不满意：”和“过程不满意：”。",
            "按两个分段分别写产物缺陷与过程缺陷。",
            reason[:220],
        )

    if len(re.findall(r"(?:产物|结果)不满意[：:]", reason)) > 1 or len(re.findall(r"过程不满意[：:]", reason)) > 1:
        add(issues, "规则 31", "不满意原因", HARD, "不满意原因分段标记重复出现。", "每个分段标记只保留一次。")

    for marker in ("产物不满意", "结果不满意", "过程不满意"):
        idx = reason.find(marker)
        if idx > 0 and reason[idx - 1] not in "。；;，,\n ":
            add(issues, "规则 29", "不满意原因", HARD, f"“{marker}”前缺少分隔标点。", "在分段标记前加句号、分号或换行。")

    vague_terms = ["效果不好", "不行", "一般", "有问题", "没完成任务", "代码有bug", "越改问题越多", "需求实现得不完整"]
    if any(term in reason for term in vague_terms) and not re.search(r"[A-Za-z0-9_/.-]+\.(?:vue|js|ts|java|py|go|md|yml|yaml|json|css|html)|接口|页面|字段|按钮|函数|方法|路由|API|报错|无法", reason):
        add(
            issues,
            "规则 9",
            "不满意原因",
            HARD,
            "不满意原因偏笼统，缺少可核查证据。",
            "写清具体文件/功能/步骤、实际表现和业务影响。",
            reason[:220],
        )

    if re.search(r"3003|100000|模型请求失败|网络波动|请求失败|服务繁忙|连接超时|connecttimeout|ssl|timeout", reason, re.I):
        add(
            issues,
            "规则 AI",
            "不满意原因",
            REVIEW,
            "不满意原因包含模型请求、网络或环境异常，规则要求不能把这类问题当作模型能力缺陷。",
            "重新执行或改写为模型本身能力造成的过程/产物问题。",
            reason[:220],
        )

    if re.search(r"Traceback|Exception|TypeError|ReferenceError|SyntaxError|npm ERR!|^\s*at\s+\w+", reason, re.I | re.M):
        add(
            issues,
            "规则 30",
            "不满意原因",
            HARD,
            "不满意原因疑似粘贴了控制台/堆栈原文。",
            "保留简短错误现象即可，不要整段粘贴日志或堆栈。",
            reason[:220],
        )

    if "从新" in reason:
        add(issues, "规则 6", "不满意原因", HARD, "存在错别字“从新”。", "改为“重新”。")

    if "死循环" in reason and "重试" not in reason:
        add(issues, "规则 19", "不满意原因", HARD, "提到“死循环”但没有说明重试行为。", "写清模型在何处反复重试、重复了什么动作。")

    if prompt and containment(reason, prompt) >= 0.55:
        add(
            issues,
            "规则 12",
            "不满意原因",
            HARD,
            "不满意原因与 User Prompt 高度重合，像是在复述需求而不是评价实际问题。",
            "改写为实际产物/过程缺陷，并补充证据。",
            f"containment={containment(reason, prompt):.2f}",
        )

    if process and not re.search(r"步骤|环节|工具|读取|搜索|修改|生成|验证|测试|运行|编译|提交|重复|遗漏|未检查|未运行|总结|规划|理解|判断|追问|定位", process):
        add(
            issues,
            "规则 15",
            "不满意原因",
            HARD,
            "过程不满意分段缺少过程或深度评价，只像是在描述产物问题。",
            "补充模型在哪个步骤做错了什么、漏了什么验证或判断。",
            process[:180],
        )

    if product and not re.search(r"[A-Za-z0-9_/.-]+\.(?:vue|js|ts|java|py|go|md|yml|yaml|json|css|html)|页面|接口|字段|按钮|路由|函数|方法|配置|样式|数据库|表|API|报错|无法|失败|缺少|未实现", product):
        add(
            issues,
            "规则 9",
            "不满意原因",
            HARD,
            "产物不满意分段缺少具体产物证据。",
            "补充文件、功能点、页面、接口、字段或可复现表现。",
            product[:180],
        )


def check_batch_context(row: dict[str, Any], history: list[dict[str, Any]], issues: list[Issue]) -> None:
    if not history:
        return

    session = session_key(text(row.get("Trae Session ID")))
    repo_commit = (text(row.get("Repo URL")), text(row.get("Commit ID")))
    prompt = text(row.get("User Prompt"))
    reason = text(row.get("不满意原因"))
    owner = text(row.get("做题人"))
    task_type = text(row.get("任务类型"))

    if session:
        for old in history:
            old_session = session_key(text(old.get("Trae Session ID")))
            if old_session and old_session == session:
                add(issues, "规则 15", "Trae Session ID", HARD, "Trae Session ID 与历史记录重复。", "新题必须使用新会话。", f"历史 Repo ID={text(old.get('Repo ID'))}")
                break

    if all(repo_commit):
        for old in history:
            if repo_commit == (text(old.get("Repo URL")), text(old.get("Commit ID"))):
                add(issues, "规则 34", "Repo URL / Commit ID", HARD, "Repo URL + Commit ID 与历史记录重复。", "每轮必须提交新的 commit，并填写对应完整哈希。", f"历史 Repo ID={text(old.get('Repo ID'))}")
                break

    if prompt:
        best_lcs = ("", 0)
        best_jaccard = ("", 0.0)
        for old in history:
            old_prompt = text(old.get("User Prompt"))
            if not old_prompt:
                continue
            common = lcs_len(prompt, old_prompt)
            sim = jaccard(prompt, old_prompt)
            if common > best_lcs[1]:
                best_lcs = (text(old.get("Repo ID")), common)
            if sim > best_jaccard[1]:
                best_jaccard = (text(old.get("Repo ID")), sim)
        if best_lcs[1] >= 40:
            add(issues, "规则 23", "User Prompt", DISCARD, "User Prompt 与历史记录存在长片段雷同。", "重写为当前任务的独立、具体需求。", f"历史 Repo ID={best_lcs[0]}，最长公共片段≈{best_lcs[1]}字")
        if best_jaccard[1] >= 0.30:
            add(issues, "规则 18", "User Prompt", DISCARD, "User Prompt 与历史记录 trigram 相似度偏高。", "避免同义改写或模板化复用。", f"历史 Repo ID={best_jaccard[0]}，Jaccard={best_jaccard[1]:.2f}")

    if reason:
        best_reason = ("", 0)
        for old in history:
            old_reason = text(old.get("不满意原因"))
            if not old_reason:
                continue
            common = lcs_len(reason, old_reason)
            if common > best_reason[1]:
                best_reason = (text(old.get("Repo ID")), common)
        if best_reason[1] >= 18:
            add(issues, "规则 37", "不满意原因", HARD, "不满意原因与历史记录连续长片段雷同。", "重新基于本轮真实过程和产物证据撰写。", f"历史 Repo ID={best_reason[0]}，最长公共片段≈{best_reason[1]}字")

    if owner:
        owned = [old for old in history if text(old.get("做题人")) == owner]
        if len(owned) >= 10:
            satisfied = sum(1 for old in owned if text(old.get("过程与产物是否满意")) == "满意")
            rate = satisfied / len(owned)
            if rate > 0.40:
                add(issues, "规则 27", "过程与产物是否满意", DISCARD, "同做题人历史满意率超过 40% 阈值。", "检查该做题人的批次分布，必要时调整或废弃超额记录。", f"{owner} 满意率={rate:.1%}")
            type_count = Counter(text(old.get("任务类型")) for old in owned if text(old.get("任务类型")))
            if task_type and type_count:
                current_share = (type_count[task_type] + 1) / (len(owned) + 1)
                if current_share > 0.55:
                    add(issues, "规则 17", "任务类型", DISCARD, "同做题人任务类型分布可能失衡。", "按批次分布要求均衡代码生成、Feature迭代、Bug修复、理解/重构等类型。", f"{task_type} 占比≈{current_share:.1%}")


def suggest_status(issues: list[Issue]) -> str:
    if any(issue.severity == DISCARD for issue in issues):
        return "建议废弃"
    if any(issue.severity == HARD for issue in issues):
        return "建议打回"
    if any(issue.severity == REVIEW for issue in issues):
        return "建议人工复核"
    return "未发现明显问题"


def check_row(row: dict[str, Any], history: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    issues: list[Issue] = []
    check_required(row, issues)
    check_repo_fields(row, issues)
    check_task_consistency(row, issues)
    check_satisfaction_and_reason(row, issues)
    check_batch_context(row, history or [], issues)
    return {
        "suggested_status": suggest_status(issues),
        "issue_count": len(issues),
        "issues": [asdict(issue) for issue in issues],
    }


def read_row_arg(args: argparse.Namespace) -> dict[str, Any] | None:
    if args.row_json:
        return json.loads(args.row_json)
    if args.row_file:
        return json.loads(Path(args.row_file).read_text(encoding="utf-8"))
    return None


def print_human(result: dict[str, Any]) -> None:
    print(f"结论：{result['suggested_status']}，发现 {result['issue_count']} 个问题")
    for index, issue in enumerate(result["issues"], 1):
        print(f"\n{index}. [{issue['severity']}] {issue['rule']} / {issue['field']}")
        print(f"   问题：{issue['message']}")
        if issue.get("suggestion"):
            print(f"   建议：{issue['suggestion']}")
        if issue.get("evidence"):
            print(f"   证据：{issue['evidence']}")


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description="Check Solo Coder spreadsheet rows.")
    parser.add_argument("--workbook", default=DEFAULT_WORKBOOK, help="Excel workbook used as history context.")
    parser.add_argument("--sheet", default=DATA_SHEET, help="Data sheet name.")
    parser.add_argument("--summary", action="store_true", help="Print workbook quality-note summary.")
    parser.add_argument("--row-json", help="A JSON object representing one new input row.")
    parser.add_argument("--row-file", help="Path to a JSON file representing one new input row.")
    parser.add_argument("--json", action="store_true", help="Print machine-readable JSON.")
    args = parser.parse_args(argv)

    history: list[dict[str, Any]] = []
    workbook_path = Path(args.workbook)
    if workbook_path.exists():
        history = load_rows(str(workbook_path), args.sheet)
    elif args.summary:
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    if args.summary:
        summary = workbook_summary(history)
        print(json.dumps(summary, ensure_ascii=False, indent=2, default=lambda value: dict(value)))

    row = read_row_arg(args)
    if row is not None:
        result = check_row(row, history)
        if args.json:
            print(json.dumps(result, ensure_ascii=False, indent=2))
        else:
            print_human(result)

    if not args.summary and row is None:
        parser.print_help()
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
